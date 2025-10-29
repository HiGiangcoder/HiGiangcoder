#!/usr/bin/env python3
import os
import sys
import subprocess
import shutil
import json
from pathlib import Path
import time
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

class TestRunner:
    def __init__(self):
        self.base_dir = Path(".")
        self.code_bug_dir = self.base_dir / "code_bug"
        self.results = {}
        
    def get_solution_files(self, problem):
        """Lấy tất cả các file solution cho một bài cụ thể"""
        problem_dir = self.code_bug_dir / f"b{problem}"
        if not problem_dir.exists():
            return []
        
        # Tìm tất cả file .cpp trong thư mục bài
        cpp_files = list(problem_dir.glob("*.cpp"))
        return cpp_files
    
    def setup_environment(self, team_dir, problem):
        """Thiết lập môi trường kiểm tra cho một bài cụ thể"""
        problem_dir = team_dir / f"b{problem}"
        
        if not problem_dir.exists():
            print(f"   ❌ Thư mục {problem_dir} không tồn tại")
            return False
            
        # Tạo thư mục tạm thời cho bài kiểm tra
        temp_dir = self.base_dir / "temp_test"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
        temp_dir.mkdir(exist_ok=True)
        
        print(f"   📁 Đang thiết lập môi trường từ {team_dir.name}/b{problem}")
        
        # Sao chép checker từ team (nếu có)
        checker_sources = [
            problem_dir / "checker.cpp",
            problem_dir / "checker"
        ]
        
        for checker_source in checker_sources:
            if checker_source.exists():
                print(f"   🔍 Tìm thấy checker: {checker_source}")
                if checker_source.suffix == ".cpp":
                    # Biên dịch checker
                    try:
                        subprocess.run([
                            "g++", "-std=c++17", "-O2", 
                            str(checker_source), "-o", str(temp_dir / "checker")
                        ], check=True, capture_output=True)
                        print("   ✅ Checker compiled successfully")
                    except subprocess.CalledProcessError as e:
                        print(f"   ❌ Lỗi biên dịch checker: {e}")
                else:
                    # Sao chép checker đã biên dịch
                    shutil.copy2(checker_source, temp_dir / "checker")
                    print("   ✅ Checker copied successfully")
                break
        
        # Sao chép validator (nếu có)
        validator_sources = [
            problem_dir / "validator.cpp", 
            problem_dir / "validator"
        ]
        
        for validator_source in validator_sources:
            if validator_source.exists():
                print(f"   🔍 Tìm thấy validator: {validator_source}")
                if validator_source.suffix == ".cpp":
                    try:
                        subprocess.run([
                            "g++", "-std=c++17", "-O2",
                            str(validator_source), "-o", str(temp_dir / "validator")
                        ], check=True, capture_output=True)
                        print("   ✅ Validator compiled successfully")
                    except subprocess.CalledProcessError as e:
                        print(f"   ❌ Lỗi biên dịch validator: {e}")
                else:
                    shutil.copy2(validator_source, temp_dir / "validator")
                    print("   ✅ Validator copied successfully")
                break
        
        # Sao chép test cases - QUAN TRỌNG: đảm bảo sao chép đúng cấu trúc
        tests_source = problem_dir / "tests"
        if tests_source.exists():
            tests_dest = temp_dir / "tests"
            print(f"   📂 Sao chép tests từ {tests_source} đến {tests_dest}")
            
            # Kiểm tra cấu trúc thư mục tests
            test_count = 0
            for root, dirs, files in os.walk(tests_source):
                if "input.in" in files:
                    test_count += 1
            
            if test_count > 0:
                # Sao chép toàn bộ cây thư mục
                shutil.copytree(tests_source, tests_dest)
                print(f"   ✅ Đã sao chép {test_count} test cases")
            else:
                print("   ⚠️ Thư mục tests tồn tại nhưng không có file input.in")
                # In cấu trúc để debug
                for root, dirs, files in os.walk(tests_source):
                    rel_path = os.path.relpath(root, tests_source)
                    print(f"      {rel_path}/: {files}")
        else:
            print("   ❌ Không tìm thấy thư mục tests")
            return False
        
        # Sao chép testlib.h (nếu cần)
        testlib_source = problem_dir / "testlib.h"
        if testlib_source.exists():
            shutil.copy2(testlib_source, temp_dir / "testlib.h")
            print("   ✅ Đã sao chép testlib.h")
            
        return True
    
    def run_test_for_solution(self, team_name, problem, solution_file):
        """Chạy kiểm tra cho một solution cụ thể"""
        team_dir = self.base_dir / team_name
        problem_code = f"b{problem}"
        
        if not team_dir.exists():
            return {"status": "missing", "message": f"Team {team_name} không tồn tại"}
        
        solution_name = solution_file.stem
        print(f"🔍 Đang kiểm tra {team_name}/{problem_code} với {solution_name}...")
        
        # Thiết lập môi trường
        if not self.setup_environment(team_dir, problem):
            return {"status": "missing", "message": f"Không tìm thấy bài {problem_code}"}
        
        # Chạy check.py
        temp_dir = self.base_dir / "temp_test"
        
        # Tạo thư mục logs
        logs_dir = self.base_dir / "logs"
        logs_dir.mkdir(exist_ok=True)
        log_file = logs_dir / f"{team_name}_b{problem}_{solution_name}.log"
        
        try:
            # Sao chép solution vào thư mục tạm
            shutil.copy2(solution_file, temp_dir / "solution.cpp")
            print(f"   ✅ Đã sao chép solution: {solution_file}")
            
            # Sao chép check.py vào thư mục tạm để đảm bảo đường dẫn đúng
            check_py_source = self.code_bug_dir / "check.py"
            shutil.copy2(check_py_source, temp_dir / "check.py")
            
            # Kiểm tra nội dung thư mục tạm trước khi chạy
            print(f"   📁 Nội dung thư mục tạm ({temp_dir}):")
            for item in temp_dir.iterdir():
                if item.is_dir():
                    print(f"      📂 {item.name}/")
                    # Liệt kê nội dung thư mục con tests
                    if item.name == "tests":
                        for test_dir in item.iterdir():
                            if test_dir.is_dir():
                                print(f"        📂 {test_dir.name}/: {list(test_dir.iterdir())}")
                else:
                    print(f"      📄 {item.name}")
            
            # Chạy check.py với thời gian chờ
            print(f"   🚀 Đang chạy check.py...")
            start_time = time.time()
            
            result = subprocess.run([
                "python3", "check.py", "solution.cpp"
            ], cwd=temp_dir, capture_output=True, text=True, timeout=300)
            
            execution_time = time.time() - start_time
            
            # Ghi log chi tiết
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"Team: {team_name}, Bài: b{problem}, Solution: {solution_name}\n")
                f.write(f"Thời gian thực thi: {execution_time:.2f}s\n")
                f.write(f"Return code: {result.returncode}\n")
                f.write("="*50 + "\n")
                f.write("STDOUT:\n")
                f.write(result.stdout)
                f.write("\n" + "="*50 + "\n")
                f.write("STDERR:\n")
                f.write(result.stderr)
            
            print(f"   ⏱️ Thời gian chạy: {execution_time:.2f}s")
            print(f"   📋 Return code: {result.returncode}")
            
            # In ra stdout và stderr để debug
            if result.stdout:
                print(f"   📝 Output: {result.stdout[-500:]}")  # In 500 ký tự cuối
            if result.stderr:
                print(f"   📝 Error: {result.stderr[-500:]}")
            
            if result.returncode == 0:
                # Phân tích kết quả từ output
                output = result.stdout + result.stderr
                
                if "Summary" in output:
                    # Trích xuất thông tin từ summary
                    lines = output.split('\n')
                    total_tests = 0
                    passed_tests = 0
                    
                    for line in lines:
                        if "Tổng số test" in line:
                            total_tests = int(line.split(":")[1].strip())
                        elif "Passed" in line and "✅" in line:
                            passed_tests = int(line.split(":")[1].strip())
                    
                    if total_tests > 0:
                        score = passed_tests / total_tests
                        return {
                            "status": "completed",
                            "passed": passed_tests,
                            "total": total_tests,
                            "score": score,
                            "execution_time": execution_time,
                            "log_file": str(log_file),
                            "solution_name": solution_name
                        }
                    else:
                        return {
                            "status": "error", 
                            "message": "Không có test cases nào được chạy",
                            "output": output[:1000],
                            "log_file": str(log_file),
                            "solution_name": solution_name
                        }
                else:
                    return {
                        "status": "error", 
                        "message": "Không tìm thấy summary trong output",
                        "output": output[:1000],
                        "log_file": str(log_file),
                        "solution_name": solution_name
                    }
            else:
                error_msg = f"Lỗi khi chạy check.py (returncode: {result.returncode})"
                print(f"   ❌ {error_msg}")
                
                return {
                    "status": "error",
                    "message": error_msg,
                    "output": result.stderr[:1000],
                    "log_file": str(log_file),
                    "solution_name": solution_name
                }
                
        except subprocess.TimeoutExpired:
            error_msg = "Chạy quá thời gian (5 phút)"
            print(f"   ⏰ {error_msg}")
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"Timeout after 300 seconds\n")
            return {
                "status": "timeout",
                "message": error_msg,
                "log_file": str(log_file),
                "solution_name": solution_name
            }
        except Exception as e:
            error_msg = f"Lỗi không mong muốn: {str(e)}"
            print(f"   💥 {error_msg}")
            with open(log_file, "w", encoding="utf-8") as f:
                f.write(f"Exception: {str(e)}\n")
            return {
                "status": "error",
                "message": error_msg,
                "log_file": str(log_file),
                "solution_name": solution_name
            }
    
    def calculate_score(self, results):
        """Tính điểm theo công thức mới"""
        # Phân loại solution và bug
        solution_result = None
        bug_results = []
        
        for res in results:
            if res.get("solution_name") == "solution":
                solution_result = res
            else:
                bug_results.append(res)
        
        # Kiểm tra solution
        solution_score = 0
        if solution_result and solution_result.get("status") == "completed":
            if solution_result.get("score", 0) == 1.0:  # Đạt điểm tối đa
                solution_score = 4
        
        # Tính điểm cho phần bug
        bug_score = 0
        q = len(bug_results)
        
        if q > 0:
            p = 0
            for bug in bug_results:
                if bug.get("status") == "completed":
                    if bug.get("score", 0) < 1.0:  # Không đạt điểm tối đa
                        p += 1
                else:
                    # Coi như không đạt điểm tối đa nếu có lỗi
                    p += 1
            
            bug_score = 2 * (p / q) ** 0.7
        
        total_score = 4 + solution_score + bug_score
        return {
            "total_score": total_score,
            "solution_score": solution_score,
            "bug_score": bug_score,
            "q": q,
            "p": p if q > 0 else 0,
            "has_solution": solution_result is not None,
            "solution_passed": solution_score > 0
        }
    
    def run_tests_for_team_problem(self, team_name, problem):
        """Chạy tất cả các solution cho một team và bài cụ thể"""
        solution_files = self.get_solution_files(problem)
        
        if not solution_files:
            return {
                "status": "no_solutions", 
                "message": f"Không tìm thấy solution nào cho bài {problem}",
                "solutions": [],
                "total_score": 0,
                "solution_score": 0,
                "bug_score": 0,
                "q": 0,
                "p": 0,
                "has_solution": False,
                "solution_passed": False
            }
        
        print(f"   📋 Tìm thấy {len(solution_files)} solution cho bài {problem}:")
        for sol in solution_files:
            print(f"      - {sol.name}")
        
        results = []
        for solution_file in solution_files:
            result = self.run_test_for_solution(team_name, problem, solution_file)
            results.append(result)
            
            # Hiển thị kết quả tức thì
            status = result.get("status", "unknown")
            if status == "completed":
                score_percent = result.get("score", 0) * 100
                passed = result.get("passed", 0)
                total = result.get("total", 0)
                time_taken = result.get("execution_time", 0)
                solution_name = result.get("solution_name", "unknown")
                print(f"      {solution_name}: ✅ {score_percent:.1f}% ({passed}/{total}) - {time_taken:.1f}s")
            elif status == "error":
                print(f"      {result.get('solution_name', 'unknown')}: ❌ Lỗi - {result.get('message', 'Unknown error')}")
            elif status == "missing":
                print(f"      {result.get('solution_name', 'unknown')}: ⚠️  {result.get('message', 'Not found')}")
            elif status == "timeout":
                print(f"      {result.get('solution_name', 'unknown')}: ⏰ Timeout")
            else:
                print(f"      {result.get('solution_name', 'unknown')}: ❓ Trạng thái không xác định")
            
            # Nghỉ giữa các solution để tránh quá tải
            time.sleep(1)
        
        # Tính điểm theo công thức mới
        score_calculation = self.calculate_score(results)
        
        return {
            "status": "completed",
            "solutions": results,
            "total_score": score_calculation["total_score"],
            "solution_score": score_calculation["solution_score"],
            "bug_score": score_calculation["bug_score"],
            "q": score_calculation["q"],
            "p": score_calculation["p"],
            "has_solution": score_calculation["has_solution"],
            "solution_passed": score_calculation["solution_passed"],
            "total_solutions": len(solution_files),
            "completed_solutions": len([r for r in results if r.get("status") == "completed"])
        }
    
    def run_all_tests(self):
        """Chạy kiểm tra cho tất cả các team và bài"""
        teams = [f"team{i:02d}" for i in range(1, 18)]
        problems = [1, 2, 3]
        
        print("🚀 Bắt đầu kiểm tra tự động cho 17 nhóm...")
        print("=" * 60)
        
        for team in teams:
            self.results[team] = {}
            print(f"\n🎯 Đang xử lý {team}...")
            
            for problem in problems:
                result = self.run_tests_for_team_problem(team, problem)
                self.results[team][f"b{problem}"] = result
                
                # Hiển thị kết quả tổng hợp cho bài
                status = result.get("status", "unknown")
                if status == "completed":
                    total_score = result.get("total_score", 0)
                    solution_score = result.get("solution_score", 0)
                    bug_score = result.get("bug_score", 0)
                    q = result.get("q", 0)
                    p = result.get("p", 0)
                    has_solution = result.get("has_solution", False)
                    solution_passed = result.get("solution_passed", False)
                    
                    print(f"   b{problem}: 📊 Điểm: {total_score:.2f}/10.0")
                    print(f"              Solution: {solution_score:.1f}/7.5 | Bug: {bug_score:.2f}/2.5")
                    print(f"              Phát hiện: {p}/{q} bug | Solution {'✅' if solution_passed else '❌'}")
                elif status == "no_solutions":
                    print(f"   b{problem}: ❌ {result.get('message', 'Không có solution')}")
                else:
                    print(f"   b{problem}: ❓ Trạng thái không xác định")
    
    def export_to_excel(self):
        """Xuất kết quả chi tiết ra file Excel"""
        print("\n📊 Đang xuất kết quả ra file Excel...")
        
        # Tạo workbook
        wb = Workbook()
        
        # Sheet tổng hợp
        ws_summary = wb.active
        ws_summary.title = "Tổng hợp"
        
        # Sheet chi tiết từng bài
        ws_b1 = wb.create_sheet("Bài 1 Chi tiết")
        ws_b2 = wb.create_sheet("Bài 2 Chi tiết") 
        ws_b3 = wb.create_sheet("Bài 3 Chi tiết")
        
        # Sheet lỗi tổng hợp
        ws_errors = wb.create_sheet("Lỗi tổng hợp")
        
        # Sheet điểm chi tiết
        ws_scores = wb.create_sheet("Điểm chi tiết")
        
        # Định nghĩa màu sắc
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # ===== SHEET TỔNG HỢP =====
        headers_summary = ["Team", "Bài 1", "Bài 2", "Bài 3", "Điểm TB", "Xếp hạng"]
        ws_summary.append(headers_summary)
        
        # Định dạng header
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Tính điểm và xếp hạng
        team_scores = {}
        for team in self.results:
            scores = []
            for problem in ["b1", "b2", "b3"]:
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    scores.append(result.get("total_score", 0))
                else:
                    scores.append(0)
            team_scores[team] = sum(scores) / len(scores) if scores else 0
        
        # Sắp xếp theo điểm
        ranked_teams = sorted(team_scores.items(), key=lambda x: x[1], reverse=True)
        ranking = {team: rank + 1 for rank, (team, score) in enumerate(ranked_teams)}
        
        # Điền dữ liệu
        for team in sorted(self.results.keys()):
            row = [team]
            for problem in ["b1", "b2", "b3"]:
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    score = result.get("total_score", 0)
                    row.append(f"{score:.2f}")
                    
                    # Tô màu theo điểm
                    if score >= 8.0:
                        cell_fill = green_fill
                    elif score >= 5.0:
                        cell_fill = yellow_fill
                    else:
                        cell_fill = red_fill
                else:
                    row.append("0.00")
                    cell_fill = red_fill
                
                # Áp dụng màu cho ô điểm
                ws_summary.cell(row=len(ws_summary['A']), column=len(row)).fill = cell_fill
            
            avg_score = team_scores[team]
            row.append(f"{avg_score:.2f}")
            row.append(ranking[team])
            ws_summary.append(row)
            
            # Tô màu cho ô xếp hạng
            if ranking[team] <= 3:
                ws_summary.cell(row=len(ws_summary['A']), column=6).fill = green_fill
        
        # ===== SHEET ĐIỂM CHI TIẾT =====
        score_headers = ["Team", "Bài", "Điểm tổng", "Điểm Solution", "Điểm Bug", "Số bug (q)", "Số bug bị phát hiện (p)", "Có solution", "Solution đạt"]
        ws_scores.append(score_headers)
        
        # Định dạng header
        for cell in ws_scores[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Điền dữ liệu điểm chi tiết
        for team in sorted(self.results.keys()):
            for problem in ["b1", "b2", "b3"]:
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    row = [
                        team, problem,
                        f"{result.get('total_score', 0):.2f}",
                        f"{result.get('solution_score', 0):.1f}",
                        f"{result.get('bug_score', 0):.2f}",
                        result.get('q', 0),
                        result.get('p', 0),
                        "Có" if result.get('has_solution') else "Không",
                        "✅" if result.get('solution_passed') else "❌"
                    ]
                    ws_scores.append(row)
        
        # ===== SHEET CHI TIẾT TỪNG BÀI =====
        detail_sheets = {"b1": ws_b1, "b2": ws_b2, "b3": ws_b3}
        detail_headers = ["Team", "Solution", "Passed", "Total", "Tỉ lệ", "Trạng thái", "Thời gian", "Ghi chú"]
        
        for problem, ws in detail_sheets.items():
            ws.append(detail_headers)
            
            # Định dạng header
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Điền dữ liệu chi tiết
            for team in sorted(self.results.keys()):
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    solutions = result.get("solutions", [])
                    for sol_result in solutions:
                        row = [team, sol_result.get("solution_name", "Unknown")]
                        
                        if sol_result.get("status") == "completed":
                            passed = sol_result.get("passed", 0)
                            total = sol_result.get("total", 0)
                            score = sol_result.get("score", 0) * 100
                            time_taken = sol_result.get("execution_time", 0)
                            
                            row.extend([
                                passed, total, f"{score:.1f}%", 
                                "HOÀN THÀNH", f"{time_taken:.2f}s", "OK"
                            ])
                            
                            # Tô màu theo tỉ lệ đúng
                            if score == 100:
                                cell_fill = green_fill
                            elif score >= 50:
                                cell_fill = yellow_fill
                            else:
                                cell_fill = red_fill
                                
                            for col in range(3, 7):  # Cột Passed đến Trạng thái
                                ws.cell(row=len(ws['A']), column=col).fill = cell_fill
                                
                        else:
                            row.extend([
                                "N/A", "N/A", "0%", 
                                "LỖI", "N/A", sol_result.get("message", "Unknown error")
                            ])
                            # Tô màu đỏ cho dòng lỗi
                            for col in range(3, 7):
                                ws.cell(row=len(ws['A']), column=col).fill = red_fill
                        
                        ws.append(row)
                else:
                    # Team không có kết quả cho bài này
                    ws.append([team, "N/A", "N/A", "N/A", "0%", "KHÔNG CÓ SOLUTION", "N/A", result.get("message", "No solutions")])
                    for col in range(3, 7):
                        ws.cell(row=len(ws['A']), column=col).fill = red_fill
        
        # ===== SHEET LỖI TỔNG HỢP =====
        error_headers = ["Team", "Bài", "Solution", "Loại lỗi", "Chi tiết"]
        ws_errors.append(error_headers)
        
        # Định dạng header
        for cell in ws_errors[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        # Thu thập tất cả lỗi
        error_count = 0
        for team in sorted(self.results.keys()):
            for problem in ["b1", "b2", "b3"]:
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    solutions = result.get("solutions", [])
                    for sol_result in solutions:
                        if sol_result.get("status") != "completed":
                            error_count += 1
                            ws_errors.append([
                                team, problem, sol_result.get("solution_name", "Unknown"),
                                sol_result.get("status", "error").upper(),
                                sol_result.get("message", "Unknown error")[:100]  # Giới hạn độ dài
                            ])
                            # Tô màu đỏ cho dòng lỗi
                            for col in range(1, 6):
                                ws_errors.cell(row=len(ws_errors['A']), column=col).fill = red_fill
                elif result.get("status") != "no_solutions":
                    error_count += 1
                    ws_errors.append([
                        team, problem, "N/A",
                        result.get("status", "error").upper(),
                        result.get("message", "Unknown error")[:100]
                    ])
                    for col in range(1, 6):
                        ws_errors.cell(row=len(ws_errors['A']), column=col).fill = red_fill
        
        # Điều chỉnh độ rộng cột
        for ws in [ws_summary, ws_scores, ws_b1, ws_b2, ws_b3, ws_errors]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Lưu file
        excel_file = "ket_qua_chi_tiet.xlsx"
        wb.save(excel_file)
        print(f"💾 Đã xuất file Excel: {excel_file}")
        print(f"   - Sheet 'Tổng hợp': Bảng điểm tổng hợp và xếp hạng")
        print(f"   - Sheet 'Điểm chi tiết': Phân tích điểm theo công thức")
        print(f"   - Sheet 'Bài X Chi tiết': Kết quả chi tiết từng solution")
        print(f"   - Sheet 'Lỗi tổng hợp': Tổng hợp {error_count} lỗi")
        
        return excel_file
    
    def generate_summary(self):
        """Tạo báo cáo tổng hợp"""
        print("\n" + "=" * 80)
        print("📊 BÁO CÁO TỔNG HỢP KẾT QUẢ")
        print("=" * 80)
        
        # Header
        header = f"{'Team':<10} {'Bài 1':<8} {'Bài 2':<8} {'Bài 3':<8} {'Điểm TB':<8} {'Xếp hạng':<10}"
        print(header)
        print("-" * 60)
        
        team_scores = {}
        
        for team in sorted(self.results.keys()):
            scores = []
            row = f"{team:<10}"
            
            for problem in ["b1", "b2", "b3"]:
                result = self.results[team].get(problem, {})
                if result.get("status") == "completed":
                    score = result.get("total_score", 0)
                    scores.append(score)
                    row += f"{score:>6.2f}  "
                else:
                    row += f"{'0.00':<8}"
                    scores.append(0)
            
            avg_score = sum(scores) / len(scores) if scores else 0
            team_scores[team] = avg_score
            row += f"{avg_score:>8.2f}  "
            print(row)
        
        # Xếp hạng
        print("-" * 60)
        print("\n🏆 BẢNG XẾP HẠNG:")
        print("-" * 40)
        
        ranked_teams = sorted(team_scores.items(), key=lambda x: x[1], reverse=True)
        
        for rank, (team, score) in enumerate(ranked_teams, 1):
            medal = ""
            if rank == 1: medal = "🥇"
            elif rank == 2: medal = "🥈" 
            elif rank == 3: medal = "🥉"
            
            print(f"{rank:2d}. {medal} {team}: {score:6.2f} điểm")
        
        # Thống kê
        print("\n📈 THỐNG KÊ:")
        print("-" * 30)
        
        total_problems = 0
        completed_problems = 0
        total_score = 0
        
        for team_results in self.results.values():
            for problem_result in team_results.values():
                if problem_result.get("status") == "completed":
                    completed_problems += 1
                    total_score += problem_result.get("total_score", 0)
                total_problems += 1
        
        if total_problems > 0:
            completion_rate = (completed_problems / total_problems) * 100
            avg_score = total_score / completed_problems if completed_problems > 0 else 0
            
            print(f"Tổng số bài kiểm tra: {total_problems}")
            print(f"Hoàn thành: {completed_problems} ({completion_rate:.1f}%)")
            print(f"Điểm trung bình: {avg_score:.2f}")
        
        # Xuất file Excel
        excel_file = self.export_to_excel()
        
        # Lưu kết quả chi tiết
        with open("detailed_results.json", "w", encoding="utf-8") as f:
            # Convert Path objects to strings for JSON serialization
            json_results = {}
            for team, problems in self.results.items():
                json_results[team] = {}
                for problem, result in problems.items():
                    json_result = result.copy()
                    if 'solutions' in json_result:
                        for sol_result in json_result['solutions']:
                            if 'log_file' in sol_result:
                                sol_result['log_file'] = str(sol_result['log_file'])
                    json_results[team][problem] = json_result
            
            json.dump(json_results, f, indent=2, ensure_ascii=False)
        
        print(f"\n💾 Kết quả chi tiết đã được lưu vào: detailed_results.json")
        print(f"📋 Logs chi tiết được lưu trong thư mục: logs/")
        print(f"📊 File Excel chi tiết: {excel_file}")
    
    def cleanup(self):
        """Dọn dẹp thư mục tạm"""
        temp_dir = self.base_dir / "temp_test"
        if temp_dir.exists():
            shutil.rmtree(temp_dir)
            print("🧹 Đã dọn dẹp thư mục tạm")

def main():
    runner = TestRunner()
    
    try:
        runner.run_all_tests()
        runner.generate_summary()
    except KeyboardInterrupt:
        print("\n⏹️ Đã dừng kiểm tra theo yêu cầu người dùng")
    except Exception as e:
        print(f"\n❌ Lỗi không mong muốn: {e}")
        import traceback
        traceback.print_exc()
    finally:
        runner.cleanup()

if __name__ == "__main__":
    main()